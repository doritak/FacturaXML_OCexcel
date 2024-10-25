import pandas as pd
from openpyxl import load_workbook

def encontrar_palabra_celda(ruta_excel, palabra):
    workbook = load_workbook(filename=ruta_excel, data_only=True)
    hoja = workbook.active
    
    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.value == palabra:
                return celda.row, celda.column
    return None, None
############# Voy a necesitar esta desde otro módulo ################
############ leer_oc(archivo, palabra)->df           ################
def leer_oc(archivo, palabra):
    inicio_fila, inicio_col = encontrar_palabra_celda(archivo, palabra)
    
    if inicio_fila is None or inicio_col is None:
        raise ValueError(f"La palabra clave '{palabra}' no se encontró en el Archivo Excel.")
    
    # Se ajusta el índice de columnas para tener el 0-based indexing
    inicio_col_letra = chr(64 + inicio_col)
    
    # Tuve que sacar este valor, porque leía muchas columnas vacías y no cortaba. Me salía error.
    # workbook = load_workbook(filename=archivo, data_only=True)
    # hoja = workbook.active
    # last_col = hoja.max_column
    ult_col = 14
    ult_col_letra = chr(64 + ult_col)
    
    # Leer los datos a partir de la celda encontrada
    df = pd.read_excel(archivo, header=inicio_fila-1, usecols=f"{inicio_col_letra}:{ult_col_letra}")
    
    # Busque la primera fila en blanco y filtre el DataFrame
    for i, row in df.iterrows():
        if row.isnull().all():
            df = df.iloc[:i]
            break
    
    return df

if __name__ == '__main__':
    archivo = 'OC nro 6791 - Patricio Lioi.xlsx'
    palabra = 'Código'
    df = leer_oc(archivo, palabra)
    print(df)

