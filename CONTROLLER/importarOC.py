import sys
import os
import pandas as pd
from openpyxl import load_workbook
# Agrega la raiz del proyecto a la ruta Python 
# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def encontrar_palabra_celda(ruta_excel, palabra):
    """
    Busca una palabra específica en un archivo de Excel y devuelve la posición de la celda que la contiene.
    Parameters:
        ruta_excel (str): La ruta del archivo de Excel.
        palabra (str): La palabra que se desea buscar en el archivo de Excel.
    Returns:
        tuple: Una tupla (fila, columna) que indica la posición de la celda que contiene la palabra.
               Si la palabra no se encuentra, devuelve (None, None).
    """
    workbook = load_workbook(filename=ruta_excel, data_only=True)
    hoja = workbook.active
    
    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.value == palabra:
                return celda.row, celda.column
    return None, None

def eliminar_columnas_vacias(df,cantidad_null):
    """
    Elimina las columnas vacías de un DataFrame con un porcentaje de nulas mayor al cantidad_nulas.
    Parameters:
        df (pandas.DataFrame): El DataFrame a procesar.
        cantidad_null (float): El porcentaje de valores nulos que determina si una columna se considera vacía.
    Returns:
        pandas.DataFrame: El DataFrame con las columnas vacías eliminadas.
    """
    umbral_porc = len(df) * (1 - cantidad_null / 100.0)
    # Eliminar columnas que no cumplan con el umbral
    df_limpio = df.dropna(axis=1, thresh=umbral_porc)
    
    return df_limpio

def leer_oc(archivo, palabra,ult_col):
    """
    Lee un archivo Excel y extrae los datos a partir de una palabra clave específica.
    Parameters:
        archivo (str): La ruta del archivo Excel a leer.
        palabra (str): La palabra clave que indica el inicio de los datos a extraer.
    Returns:
        pandas.DataFrame: Un DataFrame con los datos extraídos desde la palabra clave hasta la última columna especificada.
    Raises:
        ValueError: Si la palabra clave no se encuentra en el archivo Excel.
    """
    
    inicio_fila, inicio_col = encontrar_palabra_celda(archivo, palabra)
    
    if inicio_fila is None or inicio_col is None:
        raise ValueError(f"La palabra clave '{palabra}' no se encontró en el Archivo Excel.")
    
    # Se ajusta el índice de columnas para tener el 0-based indexing
    inicio_col_letra = chr(64 + inicio_col)
    
    # Tuve que sacar este valor, porque leía muchas columnas vacías y no cortaba. Me salía error.
    # workbook = cargar_excel_workbook(archivo)
    # hoja = workbook.active
    # ult_col = hoja.max_column
    # ult_col = 14
    ult_col_letra = chr(64 + ult_col)
    
    # Leer los datos a partir de la celda encontrada
    df = pd.read_excel(archivo, header=inicio_fila-1, usecols=f"{inicio_col_letra}:{ult_col_letra}", dtype={"Código": str})
    
    # Busque la primera fila en blanco ahí break y no siga leyendo y filtre el DataFrame
    for i, fila in df.iterrows():
        if fila.isnull().all():
            df = df.iloc[:i]
            break
    
    
    df = eliminar_columnas_vacias(df, 80)
    
    # tengo que asegurarme que la columna Código sea string
    if "Código" in df.columns:
        df["Código"] = df["Código"].astype(str)

    # adivima mejor el tipo de datos que es un objeto
    df = df.infer_objects(copy=False)
    # df = df.map(lambda x: '' if pd.isna(x) or x == 0 else x)

    
    # Reemplace los valores NaN con 0 en las columnas numéricas y convierta las columnas a números, las que son objetos quedan igual
    numeric_cols = df.select_dtypes(include=['number']).columns
    df[numeric_cols] = df[numeric_cols].fillna(0)
    df = df.fillna("")
    
    return df

if __name__ == '__main__':

    # archivo = 'XML OC Excel/../_Doc_Import/OC nro 6791.xlsx'
    # palabra = 'Código'
    # num_col = 14
    
    archivo = 'XML OC Excel/../_Doc_Import/28-10-2024_LIOI_No_Borrar.xlsx'
    palabra = 'Nro_Linea'
    num_col = 26
    
    df = leer_oc(archivo, palabra,num_col)
    print("AQUIIIIIIIIIIII")
    print(df.dtypes)
    print(df)
    
